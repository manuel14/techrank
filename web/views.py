from django.shortcuts import render, redirect
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from .models import Cliente, Tecnico
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Border, Side
from .helpers import *
from datetime import datetime
from django.http import HttpResponse
import json
import pytz
import random


@login_required(login_url='/web/login/')
def index(request):
    return render(request, 'web/main.html')


@login_required(login_url='/web/login/')
def seguimiento(request):
    clientes = Cliente.objects.exclude(estado="LI")
    for c in clientes:
        tz = pytz.timezone('America/Argentina/Buenos_Aires')
        c.fecha = c.fecha_ing.astimezone(tz)
    user_groups = json.dumps(
        list(request.user.groups.values_list('name', flat=True)))
    return render(request, 'web/seguimiento.html',
                  {'clientes': clientes, 'grupo': user_groups})


@login_required(login_url='/web/login/')
def logout_view(request):
    logout(request)
    return redirect('/web/login/')


@login_required(login_url='/web/login/')
def dato(request):
    nombre = request.POST.get("nombre", None)
    email = request.POST.get("email", None)
    telefono = request.POST.get("telefono", None)
    nodo = request.POST.get("nodo", None)
    direccion = request.POST.get("direccion", None)
    compartido = request.POST.get("compartido", None)
    id_tec = request.user.username
    tec, created = Tecnico.objects.get_or_create(
        tecnico_id=id_tec, user=request.user)
    if compartido:
        user = User.objects.get(username=compartido)
        tec_comp, created = Tecnico.objects.get_or_create(
            tecnico_id=compartido, user=user)
    else:
        tec_comp = None
    cliente = Cliente(
        nombre=nombre, email=email, direccion=direccion,
        telefono=telefono, tecnico=tec,
        nodo=nodo, tecnico_compartido=tec_comp
    )
    cliente.save()
    return redirect('/web/seguimiento')


@login_required(login_url='/web/login/')
def estados(request):
    clientes_list = []
    clientes_estados = list(request.POST.dict().items())
    # Remuevo la tupla que contiene el csrftoken
    clientes_estados = [(x, y) for (
        x, y) in clientes_estados if x != "csrfmiddlewaretoken"]
    for v in clientes_estados:
        if v[0].split("-")[0] == "clientenro":
            clientes_dic = {
                "cliente": int(v[0].split("-")[1]),
                "estado": None,
                "obs": None,
                "clientenro": v[1]
            }
        elif v[0].split("-")[0] == "estado":
            clientes_dic = {
                "cliente": int(v[0].split("-")[1]),
                "estado": v[1].upper(),
                "obs": None,
                "clientenro": None
            }
        else:
            obs = v[1].strip()
            clientes_dic = {
                "cliente": int(v[0].split("-")[1]),
                "estado": None,
                "obs": obs,
                "clientenro": None
            }
        clientes_list.append(clientes_dic)
    for c in clientes_list:
        cli_obj = Cliente.objects.get(pk=c["cliente"])
        if c["estado"]:
            cli_obj.estado = c["estado"]
            if c["estado"] == "LI":
                cli_obj.fecha_liq = datetime.now()
        if c["obs"]:
            cli_obj.observacion = c["obs"]
        if c["clientenro"]:
            cli_obj.clientenro = c["clientenro"]
        cli_obj.save()

    return redirect('/web/seguimiento')


@login_required(login_url='/web/login/')
def ranking(request):
    """Traigo todos los tecnicos que tienen cliente directo o compartido en estado
    liquidado o instalado
    """
    tecs = ranking_tecs()
    user_groups = json.dumps(
        list(request.user.groups.values_list('name', flat=True)))
    return render(request, 'web/ranking.html', {'tecnicos': tecs, "grupo": user_groups})


def excel_ranking(request):
    wb = Workbook()
    ws = wb.active
    bold_font = Font(bold=True, size=18)
    ws.title = "Instalaciones"
    titulo = "Ranking de técnicos al %s " % (
        datetime.now().strftime('%d-%m-%Y %H:%M:%S'))
    ws["A1"] = titulo
    ws["A1"].font = bold_font
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 60
    colors = ['FF000000', '0000FF', '00FF00', 'FF00FF', 'FF6600']
    rowNum = 3
    ws.cell(row=rowNum, column=1).value = "Posición"
    ws.cell(row=rowNum, column=1).font = bold_font
    ws.cell(row=rowNum, column=2).value = "Nombre"
    ws.cell(row=rowNum, column=2).font = bold_font
    ws.cell(row=rowNum, column=3).value = "Número de cliente"
    ws.cell(row=rowNum, column=3).font = bold_font
    ws.cell(row=rowNum, column=4).value = "Nombre de cliente"
    ws.cell(row=rowNum, column=4).font = bold_font
    ws.cell(row=rowNum, column=5).value = "Fecha de liquidación"
    ws.cell(row=rowNum, column=5).font = bold_font

    ws.cell(row=rowNum, column=6).value = "Cantidad de ventas"
    ws.cell(row=rowNum, column=6).font = bold_font
    ws.cell(row=rowNum, column=7).value = "Comisión"
    ws.cell(row=rowNum, column=7).font = bold_font

    tecnicos = ranking_tecs()
    for c, t in enumerate(tecnicos, start=1):
        color = random.choice(colors)
        border = Border(left=Side(border_style='medium', color=color),
                        right=Side(border_style='medium',
                                   color=color),
                        top=Side(border_style='medium',
                                 color=color),
                        bottom=Side(border_style='medium',
                                    color=color)
                        )
        rowNum += 1
        ws.cell(row=rowNum, column=1).value = c
        ws.cell(row=rowNum, column=1).border = border
        ws.cell(row=rowNum, column=2).value = t.nombre
        ws.cell(row=rowNum, column=2).border = border
        ws.cell(row=rowNum, column=6).value = t.ventas
        ws.cell(row=rowNum, column=6).border = border
        ws.cell(row=rowNum, column=7).value = t.comision
        ws.cell(row=rowNum, column=7).border = border
        clientes = t.clientes.filter(
            estado="LI") | t.clientes_comp.filter(estado="LI")
        for cli in clientes:
            ws.cell(row=rowNum, column=2).border = border
            ws.cell(row=rowNum, column=6).border = border
            ws.cell(row=rowNum, column=7).border = border
            ws.cell(row=rowNum, column=3).value = cli.clientenro
            ws.cell(row=rowNum, column=3).border = border
            ws.cell(row=rowNum, column=4).value = cli.nombre
            ws.cell(row=rowNum, column=4).border = border
            ws.cell(row=rowNum, column=5).value = cli.fecha_liq
            ws.cell(row=rowNum, column=5).border = border
            rowNum += 1

    name_file = 'Ranking_{0}.xlsx'.format(
        datetime.now().strftime('%d-%m-%Y %H:%M:%S'))
    response = HttpResponse(content=save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = 'attachment; filename={0}'.format(name_file)
    return response


def historial(request):
    filtro = request.GET.get("estado", None)
    if filtro and filtro != "Todos":
        clientes = Cliente.objects.filter(estado=filtro.upper())
    else:
        clientes = Cliente.objects.all()
    return render(request, 'web/historial.html',
                  {"clientes": clientes, "selected": filtro})


def error404(request):
    return render(request, 'web/404.html')


def error500(request):
    return render(request, 'web/500.html')
